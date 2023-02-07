from openpyxl import Workbook
wb=Workbook()
sheet=wb.active
sheet.title="Hcl"
#sheet["A1"].value=10
#sheet["B2"].value="test"
#sheet.cell(row=3,column=3).value="Hcl data"
#j=0
j=100
'''
for i in range(10,60,10):
    j+=1
    sheet.cell(row=j,column=1).value=i
wb.save("..//data//demoopenxlwrite.xlsx")

for row in sheet.iter_rows(min_row=1,max_row=5,max_col=2,min_col=2):
    for r in row:
        r.value=1
        
'''
for col in sheet.iter_cols(min_row=1,max_row=1,max_col=5,min_col=1):
    j+=100
    for d in col:
        d.value=j
wb.save("../data//demoopenxlwrite.xlsx")


